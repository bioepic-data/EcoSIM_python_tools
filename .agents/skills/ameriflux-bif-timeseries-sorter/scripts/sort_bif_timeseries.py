#!/usr/bin/env python3
"""Sort AmeriFlux BIF long-table variables into dated time-series outputs."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

REQUIRED_COLUMNS = ["SITE_ID", "GROUP_ID", "VARIABLE_GROUP", "VARIABLE", "DATAVALUE"]
DERIVED_COLUMNS = {"SERIES_CLASS", "TIMESTAMP", "TIMESTAMP_VARIABLE", "DATE_START", "DATE_END"}

TARGET_SERIES = [
    {
        "name": "total_lai",
        "label": "Total leaf area index",
        "group": "GRP_LAI",
        "variable": "LAI_TOT",
        "qualifiers": ["LAI_TYPE"],
    },
    {
        "name": "leaf_mass_per_area",
        "label": "Leaf mass per area",
        "group": "GRP_LMA",
        "variable": "LMA",
        "qualifiers": ["LMA_SPP"],
    },
    {
        "name": "aboveground_biomass",
        "label": "Above-ground crop biomass",
        "group": "GRP_AG_BIOMASS_CROP",
        "variable": "AG_BIOMASS_CROP",
        "qualifiers": ["AG_BIOMASS_CROP_ORGAN", "AG_BIOMASS_CROP_PHEN", "AG_BIOMASS_COMMENT"],
    },
    {
        "name": "canopy_height",
        "label": "Canopy height",
        "group": "GRP_HEIGHTC",
        "variable": "HEIGHTC",
        "qualifiers": ["HEIGHTC_STATISTIC"],
    },
    {
        "name": "fruit_yield",
        "label": "Fruit / grain yield",
        "group": "GRP_AG_PROD_CROP",
        "variable": "AG_PROD_CROP",
        "qualifiers": ["AG_PROD_CROP_ORGAN"],
        "filters": {"AG_PROD_CROP_ORGAN": "Fruits"},
    },
    {
        "name": "total_yield",
        "label": "Total crop production",
        "group": "GRP_AG_PROD_CROP",
        "variable": "AG_PROD_CROP",
        "qualifiers": ["AG_PROD_CROP_ORGAN"],
        "filters": {"AG_PROD_CROP_ORGAN": "Total"},
    },
]

OBSERVATION_GROUPS = {
    "GRP_AG_BIOMASS_CROP",
    "GRP_AG_PROD_CROP",
    "GRP_HEIGHTC",
    "GRP_LAI",
    "GRP_LMA",
    "GRP_BIOMASS_CHEM",
    "GRP_SOIL_CHEM",
    "GRP_SPP_O",
}

QUALIFIER_SUFFIXES = (
    "_COMMENT",
    "_DATE_UNC",
    "_METHOD",
    "_OPERATIONS",
    "_ORGAN",
    "_PHEN",
    "_PROFILE_MAX",
    "_PROFILE_MIN",
    "_SPP",
    "_STATISTIC",
    "_SURF",
    "_TYPE",
    "_UNIT",
)

QUALIFIER_NAMES = {
    "DM_COMMENT",
    "DM_DATE_UNC",
    "LAI_TYPE",
    "HEIGHTC_STATISTIC",
    "SOIL_CHEM_PROFILE_MIN",
    "SOIL_CHEM_PROFILE_MAX",
}


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def numeric_value(value: str) -> str:
    cleaned = clean_value(value).replace(",", "")
    if not cleaned:
        return ""
    try:
        return f"{float(cleaned):.12g}"
    except ValueError:
        return ""


def read_xlsx(path: Path, sheet_name: str | None) -> tuple[list[dict[str, str]], str]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl is required for Excel input. Use the bundled Codex Python runtime or install openpyxl.") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"Sheet {sheet_name!r} not found. Available sheets: {', '.join(wb.sheetnames)}")
        ws = wb[sheet_name]
    elif "AMF-BIF" in wb.sheetnames:
        ws = wb["AMF-BIF"]
        sheet_name = "AMF-BIF"
    else:
        ws = wb[wb.sheetnames[0]]
        sheet_name = ws.title

    rows = ws.iter_rows(values_only=True)
    try:
        headers = [clean_value(v) for v in next(rows)]
    except StopIteration:
        return [], sheet_name

    check_required(headers)
    out = []
    for row in rows:
        record = {headers[i]: clean_value(row[i]) if i < len(row) else "" for i in range(len(headers))}
        if any(record.get(col, "") for col in REQUIRED_COLUMNS):
            out.append(record)
    return out, sheet_name


def read_csv(path: Path) -> tuple[list[dict[str, str]], str]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        headers = reader.fieldnames or []
        check_required(headers)
        return [{k: clean_value(v) for k, v in row.items()} for row in reader], path.name


def check_required(headers: list[str]) -> None:
    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        raise SystemExit(f"Input is missing required BIF columns: {', '.join(missing)}")


def read_bif(path: Path, sheet_name: str | None) -> tuple[list[dict[str, str]], str]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(path, sheet_name)
    if suffix in {".csv", ".txt"}:
        return read_csv(path)
    raise SystemExit(f"Unsupported input suffix {path.suffix!r}; use .xlsx, .xlsm, or .csv")


def append_value(record: dict[str, str], key: str, value: str) -> None:
    if not key:
        return
    if key not in record or record[key] == "":
        record[key] = value
    elif value and value not in record[key].split(" | "):
        record[key] = f"{record[key]} | {value}"


def reconstruct_records(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: dict[tuple[str, str, str], dict[str, str]] = {}
    source_order: dict[tuple[str, str, str], int] = {}
    for index, row in enumerate(rows):
        site_id = row["SITE_ID"]
        group_id = row["GROUP_ID"]
        variable_group = row["VARIABLE_GROUP"]
        variable = row["VARIABLE"]
        value = row["DATAVALUE"]
        key = (site_id, variable_group, group_id)
        if key not in grouped:
            grouped[key] = {
                "SITE_ID": site_id,
                "VARIABLE_GROUP": variable_group,
                "GROUP_ID": group_id,
            }
            source_order[key] = index
        append_value(grouped[key], variable, value)

    records = []
    for key, record in grouped.items():
        record["_SOURCE_ORDER"] = str(source_order[key])
        records.append(record)
    return records


def is_date_field(name: str) -> bool:
    if name.endswith("_DATE_UNC"):
        return False
    return name.endswith("_DATE") or name.endswith("_DATE_START") or name.endswith("_DATE_END")


def date_fields(record: dict[str, str]) -> list[str]:
    return sorted(name for name, value in record.items() if value and is_date_field(name))


def normalize_date(value: str) -> str:
    value = clean_value(value)
    if not value:
        return ""
    if re.fullmatch(r"\d{8}", value):
        return f"{value[0:4]}-{value[4:6]}-{value[6:8]}"
    if re.fullmatch(r"\d{12}", value):
        return f"{value[0:4]}-{value[4:6]}-{value[6:8]} {value[8:10]}:{value[10:12]}"
    if re.fullmatch(r"\d{14}", value):
        return f"{value[0:4]}-{value[4:6]}-{value[6:8]} {value[8:10]}:{value[10:12]}:{value[12:14]}"
    return value


def choose_timestamp(record: dict[str, str]) -> tuple[str, str]:
    fields = date_fields(record)
    direct = [name for name in fields if name.endswith("_DATE")]
    starts = [name for name in fields if name.endswith("_DATE_START")]
    others = [name for name in fields if name not in set(direct + starts)]
    for name in direct + starts + others:
        normalized = normalize_date(record.get(name, ""))
        if normalized:
            return normalized, name
    return "", ""


def first_by_suffix(record: dict[str, str], suffix: str) -> str:
    for name in sorted(record):
        if name.endswith(suffix) and record.get(name):
            return normalize_date(record[name])
    return ""


def classify_record(record: dict[str, str]) -> str:
    variable_group = record["VARIABLE_GROUP"]
    if not choose_timestamp(record)[0]:
        return "static_metadata"
    if variable_group.startswith("GRP_DM_"):
        return "management_event"
    if variable_group in OBSERVATION_GROUPS:
        return "observation"
    return "dated_metadata"


def is_unit_field(name: str) -> bool:
    return name.endswith("_UNIT")


def is_qualifier_field(name: str) -> bool:
    if name.startswith("_") or name in REQUIRED_COLUMNS:
        return True
    if name in DERIVED_COLUMNS:
        return True
    if is_date_field(name) or is_unit_field(name):
        return True
    if name in QUALIFIER_NAMES:
        return True
    return any(name.endswith(suffix) for suffix in QUALIFIER_SUFFIXES)


def unit_for_variable(record: dict[str, str], variable: str) -> str:
    direct = f"{variable}_UNIT"
    if record.get(direct):
        return record[direct]
    for name, value in record.items():
        if name.endswith("_UNIT") and value:
            return value
    return ""


def qualifier_json(record: dict[str, str], variable: str) -> str:
    qualifiers = {}
    for name, value in sorted(record.items()):
        if not value or name.startswith("_"):
            continue
        if name in DERIVED_COLUMNS:
            continue
        if name in {"SITE_ID", "VARIABLE_GROUP", "GROUP_ID"}:
            continue
        if name == variable:
            continue
        if is_qualifier_field(name):
            qualifiers[name] = value
    return json.dumps(qualifiers, sort_keys=True)


def add_derived_fields(records: list[dict[str, str]]) -> None:
    for record in records:
        timestamp, timestamp_variable = choose_timestamp(record)
        record["SERIES_CLASS"] = classify_record(record)
        record["TIMESTAMP"] = timestamp
        record["TIMESTAMP_VARIABLE"] = timestamp_variable
        record["DATE_START"] = first_by_suffix(record, "_DATE_START")
        record["DATE_END"] = first_by_suffix(record, "_DATE_END")


def record_sort_key(record: dict[str, str]) -> tuple[str, str, str, int]:
    source_order = int(record.get("_SOURCE_ORDER") or 0)
    return (
        record.get("SERIES_CLASS", ""),
        record.get("VARIABLE_GROUP", ""),
        record.get("TIMESTAMP", ""),
        source_order,
    )


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def fieldnames_for(records: list[dict[str, str]]) -> list[str]:
    priority = [
        "SITE_ID",
        "VARIABLE_GROUP",
        "GROUP_ID",
        "SERIES_CLASS",
        "TIMESTAMP",
        "TIMESTAMP_VARIABLE",
        "DATE_START",
        "DATE_END",
    ]
    all_fields = sorted({field for record in records for field in record if not field.startswith("_")})
    return priority + [field for field in all_fields if field not in priority]


def long_rows(records: list[dict[str, str]]) -> list[dict[str, str]]:
    rows = []
    for record in records:
        if record.get("SERIES_CLASS") == "static_metadata":
            continue
        for variable, value in sorted(record.items()):
            if not value or is_qualifier_field(variable):
                continue
            rows.append(
                {
                    "SITE_ID": record["SITE_ID"],
                    "VARIABLE_GROUP": record["VARIABLE_GROUP"],
                    "GROUP_ID": record["GROUP_ID"],
                    "SERIES_CLASS": record["SERIES_CLASS"],
                    "TIMESTAMP": record.get("TIMESTAMP", ""),
                    "TIMESTAMP_VARIABLE": record.get("TIMESTAMP_VARIABLE", ""),
                    "DATE_START": record.get("DATE_START", ""),
                    "DATE_END": record.get("DATE_END", ""),
                    "VARIABLE": variable,
                    "RAW_VALUE": value,
                    "UNIT": unit_for_variable(record, variable),
                    "QUALIFIERS_JSON": qualifier_json(record, variable),
                }
            )
    return sorted(rows, key=lambda row: (row["SERIES_CLASS"], row["VARIABLE_GROUP"], row["TIMESTAMP"], row["VARIABLE"]))


def target_series_rows(records: list[dict[str, str]], spec: dict[str, Any]) -> list[dict[str, str]]:
    rows = []
    for record in records:
        if record.get("VARIABLE_GROUP") != spec["group"]:
            continue
        filters = spec.get("filters", {})
        if any(record.get(field, "").strip().lower() != str(expected).strip().lower() for field, expected in filters.items()):
            continue
        value = record.get(spec["variable"], "")
        if not value:
            continue
        row = {
            "SITE_ID": record["SITE_ID"],
            "GROUP_ID": record["GROUP_ID"],
            "TIMESTAMP": record.get("TIMESTAMP", ""),
            "TIMESTAMP_VARIABLE": record.get("TIMESTAMP_VARIABLE", ""),
            "VARIABLE_GROUP": record["VARIABLE_GROUP"],
            "VARIABLE": spec["variable"],
            "SERIES_NAME": spec["name"],
            "SERIES_LABEL": spec["label"],
            "VALUE": value,
            "VALUE_NUMERIC": numeric_value(value),
            "UNIT": unit_for_variable(record, spec["variable"]),
            "DATE_START": record.get("DATE_START", ""),
            "DATE_END": record.get("DATE_END", ""),
            "QUALIFIERS_JSON": qualifier_json(record, spec["variable"]),
        }
        for qualifier in spec["qualifiers"]:
            row[qualifier] = record.get(qualifier, "")
        rows.append(row)
    return sorted(rows, key=lambda row: (row["TIMESTAMP"], row["GROUP_ID"]))


def write_target_series(records: list[dict[str, str]], out_dir: Path, outputs: dict[str, str]) -> dict[str, Any]:
    target_dir = out_dir / "target_timeseries"
    target_dir.mkdir(parents=True, exist_ok=True)
    stale_yield_path = target_dir / "yield.csv"
    if stale_yield_path.exists():
        stale_yield_path.unlink()
    outputs["target_timeseries_dir"] = str(target_dir)
    summary = {}
    base_fields = [
        "SITE_ID",
        "GROUP_ID",
        "TIMESTAMP",
        "TIMESTAMP_VARIABLE",
        "VARIABLE_GROUP",
        "VARIABLE",
        "SERIES_NAME",
        "SERIES_LABEL",
        "VALUE",
        "VALUE_NUMERIC",
        "UNIT",
        "DATE_START",
        "DATE_END",
    ]
    for spec in TARGET_SERIES:
        rows = target_series_rows(records, spec)
        path = target_dir / f"{spec['name']}.csv"
        fieldnames = base_fields + spec["qualifiers"] + ["QUALIFIERS_JSON"]
        write_csv(path, rows, fieldnames)
        outputs[f"target_timeseries/{spec['name']}"] = str(path)
        dates = [row["TIMESTAMP"] for row in rows if row["TIMESTAMP"]]
        summary[spec["name"]] = {
            "label": spec["label"],
            "records": len(rows),
            "variable_group": spec["group"],
            "variable": spec["variable"],
            "filters": spec.get("filters", {}),
            "start": min(dates) if dates else "",
            "end": max(dates) if dates else "",
            "path": str(path),
        }
    return summary


def safe_group_name(group_name: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", group_name).strip("_") or "UNKNOWN_GROUP"


def default_output_dir(records: list[dict[str, str]], input_path: Path) -> Path:
    sites = sorted({record["SITE_ID"] for record in records if record.get("SITE_ID")})
    if len(sites) == 1:
        return Path("result") / sites[0] / "bif_timeseries"
    return Path("result") / f"{input_path.stem}_bif_timeseries"


def build_summary(
    input_path: Path,
    sheet_name: str,
    source_row_count: int,
    records: list[dict[str, str]],
    outputs: dict[str, str],
) -> dict[str, Any]:
    groups: dict[str, dict[str, Any]] = {}
    by_group = defaultdict(list)
    for record in records:
        by_group[record["VARIABLE_GROUP"]].append(record)
    for group, group_records in sorted(by_group.items()):
        classes = Counter(record["SERIES_CLASS"] for record in group_records)
        variables = sorted({field for record in group_records for field in record if field.isupper() and field not in REQUIRED_COLUMNS})
        groups[group] = {
            "records": len(group_records),
            "series_classes": dict(classes),
            "date_fields": sorted({field for record in group_records for field in date_fields(record)}),
            "variables": variables,
        }
    return {
        "input_file": str(input_path),
        "sheet": sheet_name,
        "source_rows": source_row_count,
        "records": len(records),
        "series_classes": dict(Counter(record["SERIES_CLASS"] for record in records)),
        "groups": groups,
        "outputs": outputs,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("bif_file", type=Path, help="AmeriFlux BIF .xlsx, .xlsm, or CSV file")
    parser.add_argument("--sheet", default=None, help="Excel sheet name; defaults to AMF-BIF when present")
    parser.add_argument("--out-dir", type=Path, default=None, help="Output directory; defaults to result/<SITE_ID>/bif_timeseries")
    args = parser.parse_args()

    rows, sheet_name = read_bif(args.bif_file, args.sheet)
    records = reconstruct_records(rows)
    add_derived_fields(records)
    records = sorted(records, key=record_sort_key)

    out_dir = args.out_dir or default_output_dir(records, args.bif_file)
    by_group_dir = out_dir / "by_group"
    out_dir.mkdir(parents=True, exist_ok=True)
    by_group_dir.mkdir(parents=True, exist_ok=True)

    all_records_path = out_dir / "all_records_wide.csv"
    static_path = out_dir / "static_metadata.csv"
    long_path = out_dir / "time_series_long.csv"
    summary_path = out_dir / "summary.json"

    write_csv(all_records_path, records, fieldnames_for(records))
    static_records = [record for record in records if record["SERIES_CLASS"] == "static_metadata"]
    write_csv(static_path, static_records, fieldnames_for(static_records) if static_records else fieldnames_for(records))
    long = long_rows(records)
    write_csv(
        long_path,
        long,
        [
            "SITE_ID",
            "VARIABLE_GROUP",
            "GROUP_ID",
            "SERIES_CLASS",
            "TIMESTAMP",
            "TIMESTAMP_VARIABLE",
            "DATE_START",
            "DATE_END",
            "VARIABLE",
            "RAW_VALUE",
            "UNIT",
            "QUALIFIERS_JSON",
        ],
    )

    outputs = {
        "all_records_wide": str(all_records_path),
        "time_series_long": str(long_path),
        "static_metadata": str(static_path),
        "by_group_dir": str(by_group_dir),
    }

    target_series_summary = write_target_series(records, out_dir, outputs)

    for group in sorted({record["VARIABLE_GROUP"] for record in records}):
        group_records = [record for record in records if record["VARIABLE_GROUP"] == group]
        group_path = by_group_dir / f"{safe_group_name(group)}.csv"
        write_csv(group_path, group_records, fieldnames_for(group_records))
        outputs[f"by_group/{group}"] = str(group_path)

    summary = build_summary(args.bif_file, sheet_name, len(rows), records, outputs)
    summary["target_series"] = target_series_summary
    summary_path.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    outputs["summary"] = str(summary_path)

    print(json.dumps({"output_dir": str(out_dir), **summary["series_classes"]}, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    sys.exit(main())
